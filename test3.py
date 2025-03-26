#pip install beautifulsoup4 requests openpyxl
import os

import requests #To access a webpage
from bs4 import BeautifulSoup
import openpyxl #Create a Excel file of the data
from openpyxl.reader.excel import load_workbook

#TCS
#TITAN
#RELIANCE
#HDFC BANK
#KOTAK MAHINDRA BANK

filename="output.xlsx"

# Check if the file exists
if os.path.exists(filename):
    # Load existing workbook
    excel = load_workbook(filename)
    print("Existing file loaded.")
else:
    # Create a new workbook
    excel=openpyxl.Workbook() # Create a new Excel workbook
    print("New file created.")


print(excel.sheetnames)
sheet=excel.active# Select the active worksheet
sheet.title="Stock Market"  # "Sheet" is the default name of the first worksheet
print(excel.sheetnames)
sheet.append(['Date','P/E','Average Revenue Growth','More or less'])
# Save the workbook
excel.save(filename)


# URL of the company page
url = "https://www.screener.in/company/KOTAKBANK/consolidated/"

# Set headers to mimic a real browser request
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36"
}

# Send the request
response = requests.get(url, headers=headers)

# Parse the HTML
soup = BeautifulSoup(response.text, "html.parser")

# Extract company name from <h1> tag
company_name = soup.find("h1", class_="h2 shrink-text").text.strip()

print("Company Name:", company_name)



names=soup.find_all("span", class_="name")
revenue = soup.find_all("span", class_="number")
names_final=[]
revenue_final=[]


for name in names:
    names_final.append(name.text.strip())
print("Company Name:", len(names),names_final)

for rev in revenue:
    revenue_final.append(rev.text.strip())
print("Company Revenue:", len(revenue),revenue_final)