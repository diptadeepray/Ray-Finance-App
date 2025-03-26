import requests #To access a webpage
from bs4 import BeautifulSoup
import openpyxl #Create a Excel file of the data

excel=openpyxl.Workbook()
print(excel.sheetnames)
sheet=excel.active
sheet.title="Top Rated Movies"
print(excel.sheetnames)
sheet.append(['Rank','Name','Year','Rating'])

req=requests.get("https://www.screener.in/company/TCS/consolidated/")
req.raise_for_status()
soup=BeautifulSoup(req.text,"html.parser") #This line is used to parse the contents of the website using BeautifulSoup

#span_elements = soup.select('div.comamy-ratios ul#top-ratios li.flex span')
span_elements=soup.find_all("name")
#movies=soup.find_all('flex flex-space-between') #Now movies has all the 'tr' tags that were inside the tbody tog
print(type(span_elements))

if not span_elements:
    print("ResultSet is empty.")

for movie in span_elements: #this for loop accesses every element of the list movies
    #all=movie.find('td',class_="titleColumn").get_text(strip=True)
    #all2 = movie.find('td', class_="titleColumn").get_text()
    #all3 = movie.find('td', class_="titleColumn").text
    print(movie.get_text())
    rank = "o"
    rank1="vg"

    print(rank, rank1)

    #print(movie)
    sheet.append([rank, rank1])


excel.save('TCS.xlsx')