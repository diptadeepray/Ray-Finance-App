from openpyxl import load_workbook

wb = load_workbook("example.xlsx")  # Load the file
ws = wb.active

ws['C1'] = 100
ws['C2'] = 200
wb.save("example.xlsx")


print(ws['C1'].value)  # Output: Hello
print(ws['B1'].value)  # Output: World
